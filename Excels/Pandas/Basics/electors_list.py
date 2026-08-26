import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 1. Raw Data with all 162 entries (All names formatted to UPPERCASE)
raw_data = [
    (7, "IDY2649911", "ARUN KUMAR SELAMSETTY"),
    (19, "IDY1132885", "VENKATA RAO GULLIPILLI"),
    (27, "IDY1208552", "VASU BORA"),
    (35, "IDY3161510", "SRI RAMA MURTHY KALLEPALLI"),
    (40, "IDY3177474", "SAILAJA KALLEPALLI"),
    (44, "IDY4007415", "BANGARU NAIDU KARRIMIJJI"),
    (50, "IDY1209550", "VENKATESH REDDY RANGALA"),
    (52, "IDY0709980", "RAJESHVARI PALIKEELA"),
    (59, "IDY2332922", "SURYA RAO KOTIPALLI"),
    (60, "IDY1208644", "SOBHA KOTIPALLI"),
    (62, "IDY1467001", "VENKATESWARA RAO BOTTA"),
    (66, "IDY3769478", "CHINNAMMULU POTNURU"),
    (69, "IDY0563585", "LAKSHMI K"),
    (75, "BGY1912047", "VIJAYA LAKSHMI VALLOORI"),
    (76, "BGY6497051", "VELAMALA SRINIVAS RAO"),
    (83, "IDY0904482", "RAMARAO POGIRI"),
    (90, "IDY3066438", "VENKATA GNANENDRA KUMAR OMMI"),
    (91, "IDY3595261", "RAJESWARI CHIGILIPALLI"),
    (106, "IDY2843977", "RAMESH VENNU"),
    (108, "IDY2843951", "KALPANA YENNU"),
    (114, "IDY1098680", "KUMAR POTTURU"),
    (115, "IDY1098888", "KUMARAMMA POTTURU"),
    (119, "IDY1780221", "NANDU GORRELA"),
    (122, "IDY0904607", "ASIRINAIDU DURGASI"),
    (123, "IDY0909317", "DALAMMA DURGASI"),
    (131, "IDY1551663", "NUKARAJU KOLLI"),
    (144, "IDY3820941", "CHANDRAKALA TADIVALASA"),
    (154, "IDY2149681", "LALITHA SAMPANGI"),
    (155, "IDY4080966", "REDDI VENKATA RAMANA"),
    (165, "IDY4004958", "RUPA DUKKA"),
    (168, "AWU0625814", "CHIRANJEEVI SAMPANGI"),
    (170, "XBO0976078", "NOOKA RAJU LAVETI"),
    (175, "BFX2176220", "IPPILI GAYATHRI"),
    (178, "IDY2030831", "ANURADHA OMMI"),
    (179, "IDY2986040", "LAXMI MEESALA"),
    (203, "IDY1760801", "TIRUPATI RAO DWARAPUREDD"),
    (204, "IDY0909341", "LAKSHMI KALYANI VIYYAPU"),
    (206, "IDY0610022", "CHITTIBABU OMMI"),
    (214, "IDY2711125", "VENKATA RAO BOYINA"),
    (215, "IDY2711216", "NAGAMANI BOYINA"),
    (216, "IDY1369404", "MANASA POGIRI"),
    (217, "IDY0145102", "PRASADA RAO BAYYE"),
    (218, "IDY0144071", "RUJUVELTU BAYYE"),
    (228, "IDY0909416", "RAVANAMMA KALIGOTLA"),
    (229, "IDY0909408", "PAIDAMMA BETHA"),
    (231, "IDY1553198", "SANDHYA KOYYANNA"),
    (236, "IDY2247253", "ANAND ALAJANGI"),
    (237, "IDY2546943", "BALAKRISHNA URITI"),
    (241, "IDY2729549", "SANTOSH KUMAR DEVARA"),
    (244, "IDY1865999", "SURYANARAYANA RACHARLA"),
    (257, "IDY2711448", "RAMYA BORA"),
    (263, "IDY2350262", "CHINNA BATTU"),
    (264, "IDY2350239", "RAMALAKSHMI BATTU"),
    (265, "IDY2327600", "SATYA DURGA SIVARAM BATTU"),
    (266, "IDY2350247", "ARUN KUMAR BATTU"),
    (270, "IDY2931095", "NIRMALA BORA"),
    (273, "IDY2443133", "GOVINDA RAO BANDARU"),
    (277, "IDY2443166", "JANAKI BANDARU"),
    (279, "IDY2949758", "GOWRI NAIDU GADI"),
    (281, "IDY2711067", "VEERA VENKATA RAMANA SARIKA"),
    (284, "IDY2658615", "POTHINA PREMA JYOTHI"),
    (287, "IDY2712040", "SOWJANYA LATHA KOTTAKKI"),
    (294, "IDY3829892", "SRINIVASU ADARI"),
    (299, "BGY6487102", "DASARI VARAHANARASIMHA SWAMIY"),
    (319, "IDY2436187", "GAUTAM DONKANA"),
    (320, "IDY0766048", "APPALA VENKATA SAGAR BEVARA"),
    (321, "IDY1208750", "RENUKA BEVARA"),
    (322, "IDY0631127", "KARUNA KUMARI BEVARA"),
    (324, "IDY3854221", "PREETHAM SURYA NARAYANA BEVARA"),
    (325, "IDY3855434", "LAKSHMI MANYA BEVARA"),
    (340, "IDY3325586", "ESWARA RAO MATTAPARTHI"),
    (366, "IDY3813839", "UMA SANKAR DURGAVAJJALA"),
    (367, "IDY1581900", "RAMANAMURTY YEMPADA"),
    (386, "BGY1915560", "ESTHER RUPADATTI"),
    (389, "IDY0907535", "UMA MAHESWARA RAO MALLA"),
    (401, "IDY0907626", "APPARAO JAMMU"),
    (404, "IDY0646125", "APPANNA BABU TANGETI"),
    (415, "IDY2418508", "SURI BABU CHIPPADA"),
    (416, "IDY1821447", "NARASAYAMMA CHIPPADA"),
    (429, "BGY1906825", "VARALAKSHMI KOTRADA"),
    (434, "BGY1881911", "LAKSHMANARAO KARAGANI"),
    (436, "BGY1882133", "DURGAMMA KARAGANI"),
    (441, "IDY2812600", "PYDI RAJU KANCHUMURTHY"),
    (442, "IDY2826550", "KALYANI KANCHU MURTHY"),
    (444, "IDY0400770", "RAMBABU RELLA"),
    (459, "IDY3201837", "SYAM SUNDAR JONNADA"),
    (474, "IDY2902831", "BANGARU BABU KADIMI"),
    (493, "IDY2992063", "JHANSI BHAIRAVA JYOSHULA"),
    (495, "IDY2645836", "LAKSHMI SOWMYA SAGI"),
    (498, "IDY2667608", "RAJU GOWTHU"),
    (503, "IDY0675165", "SUBHAN MOHAMMED SHA BUDDIN"),
    (506, "IDY3955606", "AJAY BHABA RAO RAMADEVU"),
    (507, "IDY0949859", "LAKSHMI RAMADEVU"),
    (516, "IDY2149418", "JAGANMOHAN PEDIREDLA"),
    (517, "IDY3792512", "YERRA VENKATA SATYAVATHI YERRA"),
    (522, "IDY0887349", "VENKATA RAMAYA YENUMULA"),
    (529, "IDY0058784", "SIMHADRI APPARAO KATTULA"),
    (532, "IDY2061802", "KRISHNAVENI KAJULURI"),
    (542, "IDY2830537", "KIRAN SERAKAM"),
    (545, "IDY0573204", "NAJUMUNNISA SYED"),
    (547, "IDY3570900", "ZARINA BEGUM ZARINA"),
    (548, "IDY3920766", "ZARINA BEGUM"),
    (549, "IDY0538306", "SAYAD RAJ KAMAL SHAIK"),
    (550, "IDY0538298", "KADHAR MOHAMMAD"),
    (552, "IDY2311439", "VENKATA RAMANAMMA SARAKAM"),
    (553, "IDY2311447", "RAJESH KUMAR SARAKAM"),
    (554, "BGY6171342", "RAMAMURTHY RAJU CHINTALAPATI"),
    (559, "IDY1208032", "PERUMALLU RAJU BHUPATHI RAJU"),
    (561, "IDY0759522", "PREMKUMARI BHUPATHIRAJU"),
    (567, "IDY1875949", "VENKATA SIMHADRI RAJU KAKARLAPUDI K K"),
    (569, "IDY3804549", "MURALI KRISHNA MOTURU"),
    (570, "IDY3804580", "RADHIKA MOTURU"),
    (584, "IDY3256047", "TANUSHA PILLA"),
    (584, "BGY6497200", "LAKSHMI DEVI KOLATI"),
    (591, "IDY1876079", "SIVA PRAKASH KONADA"),
    (594, "IDY3075744", "SAI PRAKASH KARAMPUDI"),
    (600, "IDY1724575", "YUGANDHAR BODDATI"),
    (615, "IDY0766675", "MANOJ BHIMARASETTY"),
    (620, "IDY2077148", "ANIL CHITTURI"),
    (623, "BGY1911023", "VEERA VENKATA JANARDHAN UPPADA"),
    (625, "IDY1127596", "APARNA UPPADA"),
    (634, "IDY3987690", "GOPALAKRISHNAYYA RAGHAVAPUDI"),
    (635, "IDY0641431", "AACHARI APPULABATTULA"),
    (636, "IDY0601195", "MEENA APPULABATTULA"),
    (639, "IDY0641498", "ABHISHEK APPULABATTULA"),
    (642, "IDY3943743", "DEEPTHI DASARI"),
    (643, "IDY0067884", "SAKUNTHALA SADARAM"),
    (647, "IDY2957348", "SAILENDRA KUMAR PILLA"),
    (651, "IDY0063909", "PREM CHAND JADA"),
    (652, "IDY0070847", "SADHVI MANI JADA"),
    (661, "IDY0910828", "MAHESWARARAO JAGARAPU"),
    (663, "IDY0807230", "LAKSHMANA RAO PERI"),
    (664, "IDY0807255", "PRABHAVATHI DEVI PERI"),
    (665, "IDY0766360", "ANANDA RAO KILANI"),
    (675, "IDY0896563", "NAGABHANU GUTTULA"),
    (677, "IDY0767046", "SONI MAHAMMAD"),
    (678, "IDY0761882", "MOHASIN MAHAMAD"),
    (680, "IDY1103308", "SANTHISRI DUNNA"),
    (685, "IDY1107101", "KRISHNA AYITHI REDD"),
    (688, "IDY0064386", "GANAPATHI PERICHERLA"),
    (696, "IDY0528398", "RAMA MURTHY MANTHA"),
    (698, "IDY3920683", "SATYA RAMA JYOTHI ARIPIRALA"),
    (705, "IDY0793075", "DEEPTHI SRI THOLETI"),
    (706, "IDY0428474", "V S SUBRAMANYA RAJA SHEKAR TOLETI"),
    (708, "IDY0071613", "SHOBHANA DEVI RAMINENI"),
    (713, "IDY2977833", "VENKATESH VADADA"),
    (715, "IDY0528232", "MAHESH DASARI"),
    (716, "IDY1468876", "RAJAMMA VADA PALLI"),
    (717, "IDY3327921", "SURESH DUVVI"),
    (720, "IDY1650333", "YATISH PAMIDIMUKKALA"),
    (726, "IDY0142398", "MANGAVENI MANDULA"),
    (729, "IDY3584224", "MOHINI MANDULA"),
    (733, "IDY1329440", "RAJARAO AKKARABOYINA"),
    (735, "IDY2311330", "RAMESH PONNADA"),
    (748, "IDY3255502", "SIRISHA APPALABATHULA"),
    (755, "IDY3322179", "VARUN KUMAR PUVVALA"),
    (758, "IDY3134483", "VENKATA SURYA CHAITABYA RAMBHATLA"),
    (763, "IDY3816188", "NIRAMALA RAO BALARAM MOHANTY"),
    (766, "IDY2704898", "SRINIVAS YEKKIRALA"),
    (769, "IDY2284008", "MARIYADASU KANITI"),
    (771, "IDY2283992", "KUMARI KANITI"),
    (772, "IDY0528489", "KISHORE SUSARLA"),
]

# 2. Create DataFrame and enforce Uppercase on Elector's Name
df = pd.DataFrame(raw_data, columns=["Serial No.", "EPIC No.", "Elector's Name"])
df["Elector's Name"] = df["Elector's Name"].str.upper()

# Insert sequential S.No. (1 to 162)
df.insert(0, "S.No.", range(1, len(df) + 1))

# 3. Export to Excel with professional styling
file_name = "electors_list_162_records.xlsx"
writer = pd.ExcelWriter(file_name, engine="openpyxl")
df.to_excel(writer, index=False, startrow=3, sheet_name="Electors")

workbook = writer.book
sheet = writer.sheets["Electors"]

# Styling configuration
font_family = "Segoe UI"
title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
data_font = Font(name=font_family, size=10)
bold_data_font = Font(name=font_family, size=10, bold=True)

title_fill = PatternFill(
    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
)
header_fill = PatternFill(
    start_color="2F5597", end_color="2F5597", fill_type="solid"
)
zebra_fill = PatternFill(
    start_color="F9FBFD", end_color="F9FBFD", fill_type="solid"
)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Title Block
sheet.merge_cells("A1:D1")
title_cell = sheet["A1"]
title_cell.value = "ELECTORS LIST (TOTAL 162 RECORDS)"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")
sheet.row_dimensions[1].height = 35

# Table Headers
sheet.row_dimensions[4].height = 24
for col_num in range(1, 5):
    cell = sheet.cell(row=4, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(
        horizontal="center" if col_num < 4 else "left", vertical="center"
    )

# Data Rows Styling & Zebra Striping
start_row = 5
end_row = start_row + len(df) - 1

for row_idx in range(start_row, end_row + 1):
    sheet.row_dimensions[row_idx].height = 20
    is_even = (row_idx - start_row) % 2 == 1

    for col_num in range(1, 5):
        cell = sheet.cell(row=row_idx, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        if is_even:
            cell.fill = zebra_fill

        if col_num in [1, 2]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_num == 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = bold_data_font
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Total Summary Row
total_row = end_row + 2
sheet.row_dimensions[total_row].height = 22
sheet.cell(row=total_row, column=1, value="Total Electors:").font = Font(
    name=font_family, size=10, bold=True
)
sheet.cell(row=total_row, column=1).alignment = Alignment(
    horizontal="right", vertical="center"
)
sheet.merge_cells(
    start_row=total_row, start_column=1, end_row=total_row, end_column=3
)

total_val_cell = sheet.cell(row=total_row, column=4, value=len(df))
total_val_cell.font = Font(name=font_family, size=10, bold=True)
total_val_cell.alignment = Alignment(horizontal="left", vertical="center")

# Auto-fit Column Widths
for col in sheet.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row > 1 and cell.value:
            max_len = max(max_len, len(str(cell.value)))
    sheet.column_dimensions[col_letter].width = max(max_len + 5, 12)

writer.close()
print(
    f"Successfully generated '{file_name}' with 162 capitalized records and sequential S.No.!"
)
