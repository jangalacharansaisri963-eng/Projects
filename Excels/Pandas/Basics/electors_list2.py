import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Structured Dataset (List of tuples: S.No., Serial No., EPIC No., Name)
raw_data = [
    (1, 9, "IDY0145037", "MADHUHA NALLI"),
    (2, 13, "BGY1908490", "SATYAVATHI BANDARU"),
    (3, 18, "IDY1325687", "DHARMA RAO DUSI"),
    (4, 25, "IDY3367067", "SURYANARAINRAJU GUNTURI"),
    (5, 27, "IDY2134203", "RAMKEY SAMPANGI"),
    (6, 29, "IDY2520310", "HARI KRISHNA GADI"),
    (7, 42, "IDY2520211", "RAMU ANAKAPALLI"),
    (8, 46, "IDY2536753", "SUSANNA GONTUKUPULLI"),
    (9, 47, "IDY2536688", "SURYANARAYANA VADALI"),
    (10, 51, "IDY2327402", "MOHAMMED NUSRAT UNNISA BEGUM MOHAMMED"),
    (11, 52, "IDY2327410", "MOHAMMED NAZIR AHMED MOHAMMED"),
    (12, 53, "IDY1588807", "TIRUPATHI RAO ARANGI"),
    (13, 60, "IDY0785063", "VENKATA RAMANA MURTHY GEDELA"),
    (14, 78, "IDY3338779", "GADIRAJU RAMA RAJU"),
    (15, 88, "IDY3240579", "PARVATHEESAM MEESALA"),
    (16, 93, "IDY0901355", "JEEVANKUMAR DARAM"),
    (17, 107, "XBO0067710", "AJIT DAMERA"),
    (18, 152, "IDY4011995", "SIPORAH ANGELIN SIYYADRI"),
    (19, 222, "IDY3074085", "Gadu Paydamma"),
    (20, 226, "IDY1298934", "APPANNA SABABI"),
    (21, 230, "IDY3749108", "PADMA SREE KORIMILLI"),
    (22, 232, "IDY2999092", "VENKATA RAMANA MURTHY SAMBANGI"),
    (23, 233, "IDY1338177", "HARIKA SAMBANGI"),
    (24, 234, "IDY3960812", "SRI SAI SATHWIK SALADI"),
    (25, 235, "XNC0834573", "RAJESWARI GATTINENI"),
    (26, 238, "IDY3741055", "HARI BABU GETTINENI"),
    (27, 239, "BGY7453525", "GUNTURU SRINIVASA SIVAKUMAR"),
    (28, 249, "IDY3915469", "APPALANARASAYYA YELDUTI"),
    (29, 258, "IDY3222965", "SAI SAHITHI GUNTURI"),
    (30, 261, "IDY0145185", "PRABHAVATHI DANDA"),
    (31, 266, "IDY2759934", "RAMESH SEERA"),
    (32, 267, "IDY3162476", "SAI SREENIVAS SEERA"),
    (33, 268, "IDY0149518", "RAMA KRISHNAM RAJU KALIDINDI"),
    (34, 272, "IDY1499806", "KASI RAJU ALLURI"),
    (35, 273, "IDY3832425", "Pawan Kumar Gunturu"),
    (36, 274, "IDY0149435", "LAXMINARASIMHA RAJU ALLURI"),
    (37, 275, "IDY2691087", "UMAMAHESWARI NOOKIREDDY"),
    (38, 279, "IDY3836988", "DHARMARAO POLAVARAPU"),
    (39, 282, "IDY3322328", "MADHU SUSHMANTH ACHUTA"),
    (40, 284, "IDY0620740", "VEENA KUMARI CHERUKURI"),
    (41, 287, "IDY1887001", "UDAY BHASKAR KARUMANCHI"),
    (42, 295, "IDY1452325", "SRINIVASU DASARI"),
    (43, 304, "IDY3229309", "ramana murthy sarakam"),
    (44, 305, "IDY3228061", "srilakshmi sarakam"),
    (45, 308, "IDY3125127", "Konda Saripalli"),
    (46, 309, "IDY0837716", "RAMA RAO TATINENI"),
    (47, 310, "IDY1444892", "NANGAMANI GUDAPATI"),
    (48, 311, "IDY1104405", "RAVI KUMAR SURYADEVARA"),
    (49, 314, "IDY3712957", "Swathi Priyanka Pentapati"),
    (50, 317, "IDY1956418", "ANANDARAJU PENUMETSA"),
    (51, 319, "IDY3211372", "DINAKAR VENKATESH PATHIVADA"),
    (52, 321, "IDY3074119", "Gadu Siva"),
    (53, 323, "IDY3571338", "Murali Raghupathruni"),
    (54, 324, "IDY3567443", "Suneetha"),
    (55, 330, "IDY2996981", "Badhapu Sihadri"),
    (56, 337, "NJX1174242", "C BERNADETTE NETTO"),
    (57, 338, "NJX1269670", "M JUDO ERNEST NETTO"),
    (58, 341, "IDY1452309", "RAJA RAJESWARI DEVI DASARI"),
    (59, 343, "IDY0495747", "SIYYADRI SIRAM"),
    (60, 345, "BGY6503403", "TADEPALLI ANANDA KUMAR"),
    (61, 346, "IDY3641602", "TADEPALLI BHAVANI"),
    (62, 348, "IDY2520062", "GOGADA SHANKAR RAO GOGADA"),
    (63, 349, "IDY2520070", "GOGADA SUVARAN GOGADA"),
    (64, 350, "IDY1588864", "LIKHITHA CHANDRA MEDANKULA"),
    (65, 353, "IDY1358134", "NAGA VENKATA SINDHURA AYYAGARI"),
    (66, 354, "IDY2903839", "Venkata Subba Rayudu Ayyagari"),
    (67, 356, "IDY0640136", "RATNABAI CHAMADALA"),
    (68, 359, "IDY2667475", "SARFRAAZ NAWAZ MOHAMMAD"),
    (69, 368, "IDY3212727", "VIJAYA LAKSHMI KANNEGANTI"),
    (70, 372, "BGY6503072", "RAGHAVAREDDY DWARAMPUDI"),
    (71, 374, "BGY6503288", "DWARAMPUDI VENKA TESWARAREDDY"),
    (72, 377, "IDY3832086", "AJIT BALLA"),
    (73, 382, "IDY0829663", "NIRANJAN MALLI PEDDI"),
    (74, 385, "IDY3910882", "Appa Rao Gudena"),
    (75, 388, "IDY3711082", "Nirmala Mummidivarapu"),
    (76, 389, "IDY3849361", "SIVA PRASAD ARUMULLI"),
    (77, 390, "IDY3593555", "Raja Sekhar Lingam"),
    (78, 391, "ACK0428459", "SHILONY MATHI"),
    (79, 394, "BGY3113230", "SANYASI MADHAV PANCHAGNULA"),
    (80, 395, "IDY4035499", "KRISHNA SRAVYA CHIRLA"),
    (81, 406, "IDY3712940", "Simhachalam Kottakki"),
    (82, 411, "IDY3411741", "CHINNAM NAIDU BONU"),
    (83, 414, "IDY3835493", "Rama Raghava Reddy Gandavarapu"),
    (84, 415, "IDY3831989", "Pranay Gandavarapu"),
    (85, 418, "IDY3720984", "Harsha Chowdary"),
    (86, 419, "IDY3720745", "Dhananjaya Rao Chowdari"),
    (87, 421, "IDY1923318", "VANAJAKSHI RAJANA"),
    (88, 422, "IDY3720737", "Prudhvi Chowdari"),
    (89, 430, "IDY3594702", "Ratna Kumari Mondreti"),
    (90, 432, "IDY3919743", "sundara narain rao kasisomayajula"),
    (91, 433, "IDY2552852", "USHA KAMESWARI ALIAS SUSHUMA SEELAM"),
    (92, 442, "IDY4034831", "GANGADHAR BIKKINA"),
    (93, 444, "IDY1814384", "VENKATA VEERA VARA PRASAD VASIREDDY"),
    (94, 447, "IDY3395571", "VENKATA LAKSHMI NARASIMHA RAO NEDUNURI"),
    (95, 449, "IDY2574952", "RADHA DEVI KATARI"),
    (96, 451, "IDY2574937", "KRISHNAM RAJU KATARI"),
    (97, 462, "IDY3900859", "Venkata Sai Krishna Killi"),
    (98, 469, "IDY2672509", "SAI KARTHIKEYA GOLLA"),
    (99, 470, "IDY0006320", "ESWARUDU VECHALAPU"),
    (100, 472, "IDY3992567", "RAJ PRAKASH MEDAVARAPU"),
    (101, 477, "IDY4054300", "Rajesh Varma Bhupatiraju"),
    (102, 482, "IDY3998176", "Yashwanth VNSK Chanamallu"),
    (103, 484, "IDY3810389", "MURALI CHALUMURI"),
    (104, 485, "IDY3547635", "Madhava Raju Sripathi"),
    (105, 489, "IDY4011912", "SANGAMITRA SIYYADRI"),
    (106, 491, "IDY2229269", "VENKATA SATYA VENU GOPALA SARMA BODDAPATI"),
    (107, 494, "BGY1886407", "VENKATA SUBBA REDDY KAMJULA"),
    (108, 495, "IDY2697050", "RAJESWAR RAO SAMPATHI RAO"),
    (109, 496, "IDY1662924", "PURNA CHANDRA RAO THALLURI"),
    (110, 502, "IDY4054151", "kakarlapudi kranti sree kakarlapudi"),
    (111, 506, "IDY0761965", "USHA KUMARI BASURU"),
    (112, 513, "IDY0901710", "SURYANARA YANARAJU RUDRARAJU"),
    (113, 514, "IDY2137934", "LAKSHMI RUDRARAJU"),
    (114, 515, "IDY0901850", "Nithin Varma Rudraraju"),
    (115, 516, "IDY1825837", "RAMA RAJU PENMATCHA"),
    (116, 517, "IDY2520187", "CHINASATHYAM PIRIDI"),
    (117, 519, "IDY1885773", "PARVATHI PENMATCHA"),
    (118, 520, "IDY2520195", "RAMANAMMA PIRIDI"),
    (119, 522, "IDY1885658", "NITHIN VARMA PENMATCHA"),
    (120, 526, "ACK8663817", "RAGHUPATHI THIRUMALA R CHERUKURI"),
    (121, 527, "BGY2570174", "ANURADHA CHERUKURI"),
    (122, 530, "IDY3460961", "ADI ANANDA KUMARI JAGADAM"),
    (123, 530, "ACK8766933", "SESHU RAJA KUMARI KODURU"),
    (124, 533, "IDY3844214", "Naga Jyothi Bandla"),
    (125, 534, "IDY3760865", "MADHURI SEELAM"),
    (126, 537, "UJK0381871", "JOHN VIDYA SAGAR GANGIREDLA"),
    (127, 538, "IDY1338391", "VEDAVATHI SAMBANGI"),
    (128, 539, "WJU3175833", "MADHAVI GETTINENI"),
    (129, 543, "IDY3461431", "KORUVADA JANAKI"),
    (130, 556, "IDY2991487", "Setti Apparap"),
    (131, 561, "IDY2552875", "GOWRI LATCHABOTULA"),
    (132, 569, "IDY2629129", "SESHAGIRI RAO MADASU"),
    (133, 574, "IDY0160804", "RV VRLN NEELADRI RAJU R"),
    (134, 580, "IDY3317799", "VENKATA RAVI PRASAD VEMARAJU"),
    (135, 582, "IDY1215870", "SURYA SUDARSANA RAO RAVULA"),
    (136, 589, "IDY3461639", "SAMAYAM KRISHNA PRASAD"),
    (137, 602, "IDY0626507", "TIRUPATI RAO BATANA"),
    (138, 605, "IDY2480853", "VISWANATHA VARMA KALIDINDI"),
    (139, 608, "IDY2349802", "HIMABINDU GADDIPATI"),
    (140, 614, "IDY2252955", "SANKARA NATH SISTA"),
    (141, 619, "IDY3211786", "SREE RAMULU JOSYULA"),
    (142, 628, "IDY1104439", "BHIMESWARA RAO ADAVIKOLANU"),
    (143, 629, "IDY2027655", "VENKATA NARAYANA MURTHY KANURI"),
    (144, 633, "IDY2763092", "CHHITI BABU VADDI"),
    (145, 638, "IDY1579004", "SOWMYA BHUPATHIRAJU"),
    (146, 643, "IDY1234152", "KUMAR KALIDINDI"),
    (147, 647, "IDY1224823", "LAKSHMI KUMARI KALIDINDI"),
    (148, 649, "IDY2169333", "GOUTHAM NADIMPALLI"),
    (149, 657, "IDY2923092", "SAI SRAVYA BHUPATHIRAJU"),
    (150, 663, "IDY2331395", "VIJAY SEKHAR RAJETI"),
    (151, 669, "IDY2268952", "SREENIVASARAO BATTULA"),
    (152, 670, "IDY2268960", "UDAYA SREE BATTULA"),
    (153, 672, "IDY1588906", "SANGEETA GAVARAPATI"),
    (154, 681, "IDY0368340", "VISWANADHAN DORASWAMI"),
    (155, 682, "BGY1889856", "DR INDIRA VISWANATHAN"),
    (156, 694, "IDY0795146", "BABETA GANIKA"),
    (157, 695, "IDY0724427", "NISHA GANIKA"),
    (158, 697, "IDY1129832", "VENU KHEMKA"),
    (159, 699, "IDY1130699", "PREETI KHEMKA"),
    (160, 704, "IDY2229293", "VENKATA RAMANA RAO BODDAPATI"),
    (161, 705, "IDY2229277", "USHASREE VALLURI"),
    (162, 714, "IDY3141207", "PRAVALLIKA GUDLA"),
    (163, 715, "IDY3143112", "LEELA LAHARI GUDLA"),
    (164, 716, "IDY2898385", "SURYANARAYANA ALLURI"),
    (165, 722, "IDY2923381", "CHANDRA SEKHAR PATIBANDLA"),
    (166, 723, "IDY2914901", "RENUKA DEVI MADDULA"),
    (167, 724, "IDY0768580", "ANIL KUMAR NETHINOORIE"),
    (168, 726, "IDY3211679", "VENKATA SUBBARAO TADINADA"),
    (169, 736, "IDY3793890", "MADHU SUSHMANTH ACHUTA"),
    (170, 738, "IDY2672517", "NAGA SOUMYA GOLLA"),
    (171, 740, "IDY2552883", "SHIVAJI KOTA"),
    (172, 742, "IDY3327582", "BALAGANGADHARA TILAK JANGAREDDY"),
    (173, 751, "IDY4006417", "simma vydehi"),
]

# 2. Process data with Pandas & Capitalize all names cleanly (.upper())
formatted_rows = [
    [s_no, serial_no, epic_no, name.upper()]
    for s_no, serial_no, epic_no, name in raw_data
]

df = pd.DataFrame(
    formatted_rows, columns=["S.No.", "Serial No.", "EPIC No.", "Elector's Name"]
)

# 3. Build Excel with openpyxl (Professional UI Styling)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Electors List"
ws.views.sheetView[0].showGridLines = True

# Colors & Fills (Corporate Navy Palette)
HEADER_FILL = PatternFill(
    start_color="1B365D", end_color="1B365D", fill_type="solid"
)
ZEBRA_FILL = PatternFill(
    start_color="F4F6F9", end_color="F4F6F9", fill_type="solid"
)
WHITE_FILL = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
)

# Typography
font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Segoe UI", size=10, color="333333")

# Borders
thin_border_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(
    left=thin_border_side,
    right=thin_border_side,
    top=thin_border_side,
    bottom=thin_border_side,
)
border_header = Border(
    left=thin_border_side,
    right=thin_border_side,
    top=thin_border_side,
    bottom=Side(border_style="medium", color="0B1D38"),
)

# Title Block Layout
ws.merge_cells("A1:D1")
ws["A1"] = "OFFICIAL ELECTORS LIST"
ws["A1"].font = font_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:D2")
ws["A2"] = f"Dataset Report | Total Records: {len(df)}"
ws["A2"].font = font_subtitle
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 10  # Spacer row

# Headers Row
for col_num, header in enumerate(df.columns, 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.font = font_header
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(
        horizontal="center" if col_num <= 3 else "left", vertical="center"
    )
    cell.border = border_header
ws.row_dimensions[4].height = 24

# Populate Rows with Zebra Striping
for r_idx, row in enumerate(df.values, start=5):
    ws.row_dimensions[r_idx].height = 20
    row_fill = ZEBRA_FILL if (r_idx % 2 == 0) else WHITE_FILL

    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = font_data
        cell.fill = row_fill
        cell.border = border_cell

        # Alignments & Formatting
        if c_idx in [1, 2]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "#,##0"
        elif c_idx == 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Custom Column Widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 45

# Freeze Header Row
ws.freeze_panes = "A5"

# Save File
output_file = "Electors_List_Report.xlsx"
wb.save(output_file)
print(f"Dataset successfully compiled and saved to '{output_file}'!")
