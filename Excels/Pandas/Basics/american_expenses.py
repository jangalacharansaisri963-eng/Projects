import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_expense_sheet():
    # 1. Prepare data categorized by typical US monthly expenses
    categories = [
        "Housing", "Housing", "Housing", "Housing",
        "Transportation", "Transportation", "Transportation",
        "Food", "Food",
        "Utilities", "Utilities", "Utilities",
        "Insurance & Debt", "Insurance & Debt", "Insurance & Debt",
        "Savings & Discretionary", "Savings & Discretionary", "Savings & Discretionary"
    ]
    
    items = [
        "Rent / Mortgage", "Property Tax", "Homeowners / Renters Insurance", "Maintenance & Repairs",
        "Car Payment", "Auto Insurance", "Gas & Public Transit",
        "Groceries", "Dining Out & Coffee",
        "Electricity & Gas", "Water, Trash & Sewer", "Internet & Phone",
        "Health Insurance", "Student Loan / Personal Debt", "Life Insurance",
        "Emergency Fund / Savings", "Subscriptions & Entertainment", "Shopping & Personal Care"
    ]
    
    avg_costs = [
        1800.0, 300.0, 100.0, 150.0,
        550.0, 150.0, 200.0,
        520.0, 300.0,
        200.0, 80.0, 150.0,
        250.0, 350.0, 60.0,
        400.0, 100.0, 200.0
    ]

    df = pd.DataFrame({
        "Category": categories,
        "Expense Item": items,
        "Monthly Budget ($)": avg_costs
    })

    file_name = "average_american_monthly_expenses.xlsx"

    # 2. Write to Excel using openpyxl for deep styling control
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Monthly Budget")

    # 3. Load workbook and apply styling
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Monthly Budget"]

    # Colors & Styles Definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Classic Navy
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft light blue
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # Style Header Row
    for col_num in range(1, 4):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style Data Rows & Currency Formatting
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, 4):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            # Align text left, numbers right
            if col_num == 3:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Add Summary / Total Row at the bottom
    total_row_num = len(df) + 2
    ws.cell(row=total_row_num, column=1, value="Total Monthly Expenses").font = bold_font
    ws.cell(row=total_row_num, column=3, value=f"=SUM(C2:C{len(df)+1})")
    
    # Format Total Row
    for col_num in range(1, 4):
        cell = ws.cell(row=total_row_num, column=col_num)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = total_border
        if col_num == 3:
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Auto-adjust Column Widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # Format string check for totals formula length representation
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "$5,000.00" 
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # Save finalized styled workbook
    wb.save(file_name)
    print(f"Successfully generated and styled '{file_name}'!")

if __name__ == "__main__":
    generate_expense_sheet()
  
